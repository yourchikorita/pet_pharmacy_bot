# -*- coding: utf-8 -*-
"""
Created on Sat Sep 21 19:02:42 2019

@author: EJ
"""

# -*- coding: utf-8 -*-

import requests
from flask import Flask, request, Response
from pet_pharmacy import pet_pharm_api
from openpyxl import load_workbook

EXCEL_FILE_NAME = 'Database.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)
tuto_db = db['fv']

API_KEY = '916473331:AAHDtuAaRcGM8FlPRFUYDGRzY6mgyDWacuo'

app = Flask(__name__)




def write_with_index(userLocList):

    tuto_db['A1'].value = userLocList
    db.save(EXCEL_FILE_NAME)
   
 
def write_user_choice_num(user_choice_num):

    tuto_db['A5'].value = user_choice_num
    db.save(EXCEL_FILE_NAME)
    print('엑셀파일에 저장되었따 유저가 선택한 번호가   ')
    
def write_final(show):

    tuto_db['A10'].value = show
    db.save(EXCEL_FILE_NAME)
    print('엑셀파일에 저장되었따 유저가 선택한 번호가   ')
    
def write_title_list(userTitle):

    tuto_db['A3'].value = userTitle
    db.save(EXCEL_FILE_NAME)
    print('엑셀파일에 저장됨!')
    
    
def write_detail_list(user_festival_list):

    tuto_db['A4'].value = user_festival_list
    db.save(EXCEL_FILE_NAME)
    #print('엑셀파일에 저장!유저가 지역 선택한 약국리스트 데이타가!')
    print('엑셀파일에 A4저장됨!')    
    
    
def read_with_index(loc):
    read_result = tuto_db[loc].value
    print('엑셀을 읽어들어옴!')
    return read_result

def read_with_sm_pet_hospital(user_pick_name='차오름동물병원'):
    sm_pet=load_workbook(filename="small_pet_database.xlsx", data_only=True)
    sm_db=sm_pet['Sheet1']
    all_values = []
    for row in sm_db.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)
    
    #서울특별시내 소동물병원 목록.
    sm_title_list=''
    for item in all_values[1:]:
      i=1
      sm_title_list=sm_title_list+item[i-1] +' ('+item[i]+')'+'\n'
      i=i+1
      
    #선택한 병원 디테일정보
    detail_info=''
    detail_line=''
    for item in all_values:
        if user_pick_name in item[0]: 
            #detail_line=all_values[0][0]+':'+item[0]+'\n'+all_values[0][1]+':'+item[1]+'\n'+all_values[0][2]+':'+item[2]
            #print(detail_line)
            detail_info=detail_info+str(item)
            detail_info_list=detail_info.split(',')
            
            name=all_values[0][0]+'-'+detail_info_list[0][1:]+'\n'
            gu=all_values[0][1]+'-'+detail_info_list[1]+'\n'
            time=all_values[0][2]+'-'+detail_info_list[2]+'\n'
            time_weekend=all_values[0][3]+'-'+detail_info_list[3]+'\n'
            tel=all_values[0][4]+'-'+detail_info_list[4]+'\n'
            addr=all_values[0][5]+'-'+detail_info_list[5]+'\n'
            ref=all_values[0][6]+detail_info_list[6][:-2]
            result = name+gu+time+time_weekend+tel+addr+ref
            
    return result,sm_title_list
    
def parse_message(data):
    chat_id = data['message']['chat']['id']
    msg = data['message']['text']
    return chat_id, msg

def pick_list_back(pick_list):
    return pick_list

def send_message_map_url(chat_id,text):
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)  
    google_map_url='https://www.google.com/maps/search/?api=1&query='
    google_map_user_url=(google_map_url+text).replace(" ","")
    params = {'chat_id':chat_id, 'text':google_map_user_url}
    requests.post(url, json=params)
       
def send_message(chat_id, text='bla-bla-bla'):

    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)   #sendMessage
    keyboard = {                                        # Keyboard 형식
            'keyboard':[[{
                    'text': '동물약국 찾기'
                        },
                    {'text': '소동물병원 찾기'
                        }],[
                                
                    {'text': '반려견 안전관리 및 과태료'
                        }]
                    ],
            'one_time_keyboard' : True
            }
    
    
    if  text[:5] == "지역검색!":
        user_loc=text[5:] # ex)노원구 
        print('------------사용자가 지역검색-------:',user_loc)
        userTitle,userLocList =  pet_pharm_api(user_loc) #함수호출
        write_title_list(userTitle)#엑셀에 작성하는 코드 A3에 타이틀리스트 저장  
        write_detail_list(userLocList)#엑셀에 작성하는 코드 A4에 디테일리스트 저장  
        read_title_A3=read_with_index('A3')
        result=user_loc+'에 있는 동물약국 목록 입니다!!! 번호 하나 선택 바람 입력방식=[더보기!+번호] 예)더보기!3번' +'\n'+'\n'+read_title_A3      
        params = {'chat_id':chat_id, 'text': result}
        requests.post(url, json=params)
    elif text == '소동물병원 찾기':
         detail_info,sm_hospital_total_list = read_with_sm_pet_hospital()
         params = {'chat_id':chat_id, 'text': '서울특별시 소동물병원 리스트 입니다. 상세정보를 원하시는 병원명을 입력하세요.\n'+sm_hospital_total_list}
         requests.post(url, json=params)
    elif text == '반려견 안전관리 및 과태료':     
         safe_rule = tuto_db['A20':'A42']
         safe_fine=''
         for row in safe_rule:
                for cell in row:
                    safe_fine=safe_fine+str(cell.value)+'\n'
         safe_fine_str=str(safe_fine)           
         params = {'chat_id':chat_id, 'text': safe_fine_str}
         requests.post(url, json=params)     
         
    elif text[-2:]=='병원':
         user_hopital_name=text #병원이름 들어감
         detail_info,sm_hospital_total_list=read_with_sm_pet_hospital(user_hopital_name)
         params = {'chat_id':chat_id, 'text': detail_info}
         requests.post(url, json=params)
    elif text[:4]=='더보기!':
        num=text[4:]# 예)2번
  
        #엑셀에서 엔터로 구분해서 읽으들어온다.str->list
        db_a4_str=read_with_index('A4') #A4셀에 저장해두었던 상세정보 읽어온다.
        db_a4_list=db_a4_str.split('\n')#엔터로 구분해서 list로 변환.
      
        #user가 선택한 번호를 db_a4_list에서 찾아서 디테일정보를 가져온다.     
        for item in db_a4_list:
            if num in item:
                detail=item
                    
        params = {'chat_id':chat_id, 'text': detail}
        requests.post(url, json=params)
        
        #구글 지도 유알엘 출력해주는 센더함수 
        sep_addr=detail.split('주소:')
        sep_slash=sep_addr[1].split('/')
        final_addr=sep_slash[0]
        send_message_map_url(chat_id,text=final_addr) 
        
    elif text=="동물약국 찾기":
        params = {'chat_id':chat_id, 'text': '검색을 원하는 지역을 말해줘! 입력방식=[지역검색!+검색을 원하는 지역구] 예)지역검색!노원구'}
        requests.post(url, json=params)
        
    elif text=='아니요':
        
        params = {'chat_id':chat_id, 'text': '아니라니...대화가 종료되었다네... 날 다시 부르려면 [나와라]라고 입력하시오..'}
        requests.post(url, json=params)
    else:
        params = {'chat_id':chat_id, 'text': '동물약국찾기 힘들다면.... . ', 'reply_markup' : keyboard}
        requests.post(url, json=params)

    
    return 0


    
# 경로 설정, URL 설정
@app.route('/', methods=['POST', 'GET'])
def index():
    if request.method == 'POST':
        message = request.get_json()           
        chat_id, msg = parse_message(message)
        send_message(chat_id, msg)
        return Response('ok', status=200)
    else:
        return 'Hello World!'



if __name__ == '__main__':
    app.run(port = 5000)